
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from subprocess import CREATE_NO_WINDOW
serv=Service()
serv.creation_flags = CREATE_NO_WINDOW   
driver = webdriver.Chrome(service = serv)

import time, datetime
import os#폴더 만드는 모듈
import openpyxl

now=str(datetime.datetime.now())[:16]#지금 시간을 15번째 칸까지 자르기
folderName=now.replace(":","_")#위에 있는 now에서 :을 _로 바꾸기
os.mkdir(folderName)#folderName을 폴더 이름으로 해서 폴더 만들기

key_word=["캠핑용품","캠핑장 예약"]

wb=openpyxl.Workbook()#액셀 열기

for i in range(len(key_word)):#key_word리스트의 길이만큼 반복
    ws=wb.create_sheet()#워크시트 만들기
    ws.title=key_word[i]#워크시트 이름 정하기
    ws.column_dimensions["A"].width=90#A열 너비 조정
    
    url= "https://search.naver.com/search.naver?where=news&ie=utf8&sm=nws_hty&query=" + key_word[i]
    driver.get(url)
    time.sleep(2)

    list_news=driver.find_element("class name","list_news")
    news_boxes=list_news.find_elements("class name","bx")

    for j in range(len(news_boxes)):
        driver.execute_script("arguments[0].scrollIntoView(true);", news_boxes[j])#스크롤
        file=f"{folderName}/{i+1}_{key_word}-{j+1}.png"#파일 이름 정하기
        news_boxes[j].screenshot(file)#스크린샷해서 위 변수 값을 이름으로 저장
        
        ws.row_dimensions[j+1].height=100#행 높이 조절
        img=openpyxl.drawing.image.Image(file)#스크린샷의 이미지 형식 변환(엑셀 삽입 가능으로)
        ws.add_image(img, f'A{j+1}')#A{j+1} 셀에 이미지 삽입
        
        title=news_boxes[j].find_element("class name","news_tit")#뉴스 제목을 title로 저장
        print(j+1, title.text)#뉴스 제목들을 프린트
        
        link=title.get_attribute("href")#링크 저장
        ws[f'B{j+1}'].value="기사링크"
        ws[f'B{j+1}'].hyperlink=link#B{j+1} 셀에 링크 삽입

    print()#한 줄 띄우기

wb.remove(wb["Sheet"])#시트 삭제
wb.save(f"{folderName}/{folderName}_결과.xlsx")#엑셀 저장
